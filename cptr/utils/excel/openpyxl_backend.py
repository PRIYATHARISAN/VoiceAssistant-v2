"""OpenPyXL Excel backend implementation for .xlsx file operations."""

from __future__ import annotations

import logging
import os
import re
import shutil
from pathlib import Path
from typing import Any, List, Optional

try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, column_index_from_string

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    openpyxl = None  # type: ignore[assignment]


from cptr.utils.excel.backend_base import ExcelBackend, ExcelResult

logger = logging.getLogger(__name__)


def _parse_cell_coord(cell_str: str) -> tuple[int, int]:
    """Convert 'A1' -> (row 1, col 1)."""
    match = re.match(r"^([A-Za-z]+)([0-9]+)$", cell_str.strip())
    if not match:
        raise ValueError(f"Invalid cell coordinate: '{cell_str}'")
    col_str, row_str = match.groups()
    return int(row_str), column_index_from_string(col_str.upper())


def _parse_color(color_hex: str) -> str:
    """Normalize hex color string (e.g., 'FF0000', '#FF0000', 'yellow')."""
    color = color_hex.strip().lstrip("#")
    named_colors = {
        "red": "FF0000",
        "green": "00FF00",
        "blue": "0000FF",
        "yellow": "FFFF00",
        "orange": "FFA500",
        "purple": "800080",
        "gray": "808080",
        "light_yellow": "FFFFE0",
        "light_green": "E0FFE0",
        "light_red": "FFE0E0",
    }
    return named_colors.get(color.lower(), color.upper())


def is_openpyxl_available() -> bool:
    """Check whether the openpyxl library is installed and usable."""
    return _OPENPYXL_AVAILABLE


class OpenPyXLBackend(ExcelBackend):
    """File-based Excel backend using openpyxl."""

    def __init__(self, file_path: str | None = None):
        if not _OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for file-based Excel operations. "
                "Install it with: pip install openpyxl"
            )
        self.file_path = file_path or ""
        self.wb: openpyxl.Workbook | None = None
        self.active_sheet_name: str = ""

        if self.file_path and os.path.exists(self.file_path):
            self.open_workbook(self.file_path)

    @property
    def backend_type(self) -> str:
        return "openpyxl"

    @property
    def is_live_mode(self) -> bool:
        return False

    def _get_sheet(self, sheet_name: str | None = None) -> openpyxl.worksheet.worksheet.Worksheet:
        if self.wb is None:
            raise RuntimeError("No workbook is open.")
        name = sheet_name or self.active_sheet_name or self.wb.active.title
        if name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{name}' not found in workbook.")
        return self.wb[name]

    def open_workbook(self, file_path: str = "") -> ExcelResult:
        if not file_path:
            if self.wb is None:
                self.wb = openpyxl.Workbook()
                self.active_sheet_name = self.wb.active.title
            return ExcelResult(
                success=True,
                operation="open_workbook",
                workbook=os.path.basename(self.file_path) if self.file_path else "Workbook1.xlsx",
                sheet=self.active_sheet_name,
                message="Opened in-memory Excel workbook.",
                data={"sheets": self.wb.sheetnames, "active_sheet": self.active_sheet_name},
            )
        if not os.path.exists(file_path):
            return ExcelResult(
                success=False,
                operation="open_workbook",
                workbook=file_path,
                message=f"File not found: '{file_path}'",
            )
        try:
            self.file_path = file_path
            self.wb = openpyxl.load_workbook(file_path, data_only=False)
            self.active_sheet_name = self.wb.active.title
            return ExcelResult(
                success=True,
                operation="open_workbook",
                workbook=os.path.basename(file_path),
                sheet=self.active_sheet_name,
                message=f"Workbook '{os.path.basename(file_path)}' opened successfully ({len(self.wb.sheetnames)} sheets).",
                data={"sheets": self.wb.sheetnames, "active_sheet": self.active_sheet_name},
            )
        except Exception as exc:
            return ExcelResult(
                success=False,
                operation="open_workbook",
                workbook=file_path,
                message=f"Error opening workbook: {exc}",
            )

    def create_workbook(self, file_path: str = "") -> ExcelResult:
        try:
            self.file_path = file_path
            self.wb = openpyxl.Workbook()
            self.active_sheet_name = self.wb.active.title
            if file_path:
                self.wb.save(file_path)
            return ExcelResult(
                success=True,
                operation="create_workbook",
                workbook=os.path.basename(file_path) if file_path else "Workbook1.xlsx",
                sheet=self.active_sheet_name,
                message=f"Created new workbook at '{file_path}'." if file_path else "Created new in-memory workbook.",
            )
        except Exception as exc:
            return ExcelResult(
                success=False,
                operation="create_workbook",
                workbook=file_path,
                message=f"Error creating workbook: {exc}",
            )

    def get_workbook_info(self) -> ExcelResult:
        if self.wb is None:
            return ExcelResult(success=False, operation="get_workbook_info", message="No workbook open.")
        return ExcelResult(
            success=True,
            operation="get_workbook_info",
            workbook=os.path.basename(self.file_path),
            sheet=self.active_sheet_name,
            data={
                "file_path": self.file_path,
                "sheets": self.wb.sheetnames,
                "active_sheet": self.active_sheet_name,
                "total_sheets": len(self.wb.sheetnames),
            },
        )

    def save_workbook(self, target_path: str | None = None) -> ExcelResult:
        if self.wb is None:
            return ExcelResult(success=False, operation="save_workbook", message="No workbook open.")
        save_to = target_path or self.file_path
        try:
            self.wb.save(save_to)
            self.file_path = save_to
            return ExcelResult(
                success=True,
                operation="save_workbook",
                workbook=os.path.basename(save_to),
                message=f"Workbook saved successfully to '{save_to}'.",
            )
        except Exception as exc:
            return ExcelResult(
                success=False,
                operation="save_workbook",
                workbook=save_to,
                message=f"Error saving workbook: {exc}",
            )

    def close_workbook(self) -> ExcelResult:
        if self.wb is not None:
            try:
                self.wb.close()
            except Exception:
                pass
            self.wb = None
            self.active_sheet_name = ""
        return ExcelResult(success=True, operation="close_workbook", message="Workbook closed.")

    def list_sheets(self) -> ExcelResult:
        if self.wb is None:
            return ExcelResult(success=False, operation="list_sheets", message="No workbook open.")
        return ExcelResult(
            success=True,
            operation="list_sheets",
            workbook=os.path.basename(self.file_path),
            data={"sheets": self.wb.sheetnames, "active_sheet": self.active_sheet_name},
        )

    def create_sheet(self, sheet_name: str, index: int | None = None) -> ExcelResult:
        if self.wb is None:
            return ExcelResult(success=False, operation="create_sheet", message="No workbook open.")
        if sheet_name in self.wb.sheetnames:
            return ExcelResult(
                success=False,
                operation="create_sheet",
                message=f"Sheet '{sheet_name}' already exists.",
            )
        try:
            ws = self.wb.create_sheet(title=sheet_name, index=index)
            self.active_sheet_name = ws.title
            return ExcelResult(
                success=True,
                operation="create_sheet",
                workbook=os.path.basename(self.file_path),
                sheet=sheet_name,
                message=f"Created worksheet '{sheet_name}'.",
            )
        except Exception as exc:
            return ExcelResult(
                success=False,
                operation="create_sheet",
                message=f"Error creating sheet: {exc}",
            )

    def delete_sheet(self, sheet_name: str) -> ExcelResult:
        if self.wb is None:
            return ExcelResult(success=False, operation="delete_sheet", message="No workbook open.")
        if len(self.wb.sheetnames) <= 1:
            return ExcelResult(
                success=False,
                operation="delete_sheet",
                message="Cannot delete the only sheet in the workbook.",
            )
        if sheet_name not in self.wb.sheetnames:
            return ExcelResult(
                success=False,
                operation="delete_sheet",
                message=f"Sheet '{sheet_name}' not found.",
            )
        try:
            del self.wb[sheet_name]
            self.active_sheet_name = self.wb.sheetnames[0]
            return ExcelResult(
                success=True,
                operation="delete_sheet",
                workbook=os.path.basename(self.file_path),
                sheet=self.active_sheet_name,
                message=f"Deleted sheet '{sheet_name}'.",
            )
        except Exception as exc:
            return ExcelResult(
                success=False,
                operation="delete_sheet",
                message=f"Error deleting sheet: {exc}",
            )

    def rename_sheet(self, old_name: str, new_name: str) -> ExcelResult:
        if self.wb is None:
            return ExcelResult(success=False, operation="rename_sheet", message="No workbook open.")
        if old_name not in self.wb.sheetnames:
            return ExcelResult(
                success=False,
                operation="rename_sheet",
                message=f"Sheet '{old_name}' not found.",
            )
        try:
            ws = self.wb[old_name]
            ws.title = new_name
            if self.active_sheet_name == old_name:
                self.active_sheet_name = new_name
            return ExcelResult(
                success=True,
                operation="rename_sheet",
                workbook=os.path.basename(self.file_path),
                sheet=new_name,
                message=f"Renamed sheet '{old_name}' to '{new_name}'.",
            )
        except Exception as exc:
            return ExcelResult(
                success=False,
                operation="rename_sheet",
                message=f"Error renaming sheet: {exc}",
            )

    def get_sheet_info(self, sheet_name: str | None = None) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            return ExcelResult(
                success=True,
                operation="get_sheet_info",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
                data={
                    "name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "used_range": f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
                },
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="get_sheet_info", message=str(exc))

    def read_range(
        self,
        cell_range: str | None = None,
        sheet_name: str | None = None,
        max_rows: int = 100,
        offset_row: int = 1,
    ) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            max_r = min(max_rows, 5000)  # Safe pagination cap

            if not cell_range:
                start_row = offset_row
                end_row = min(ws.max_row, start_row + max_r - 1)
                start_col = 1
                end_col = max(ws.max_column, 1)
                cell_range = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
            
            rows_data = []
            selection = ws[cell_range]
            # openpyxl returns a single Cell for single-cell refs like "C2",
            # a tuple of Cells for a single-row range like "A1:C1",
            # or a tuple of tuples for multi-row ranges like "A1:C4".
            from openpyxl.cell.cell import Cell
            if isinstance(selection, Cell):
                rows_data = [[selection.value]]
            elif isinstance(selection, tuple) and selection and isinstance(selection[0], Cell):
                # Single row of cells (e.g., "A1:C1")
                rows_data = [[cell.value for cell in selection]]
            else:
                for row in selection:
                    if isinstance(row, tuple):
                        r_vals = [cell.value for cell in row]
                    else:
                        r_vals = [row.value]
                    rows_data.append(r_vals)

            headers = rows_data[0] if rows_data else []
            return ExcelResult(
                success=True,
                operation="read_range",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell_range,
                message=f"Read {len(rows_data)} rows from '{cell_range}'.",
                data={
                    "range": cell_range,
                    "headers": headers,
                    "rows": rows_data,
                    "total_rows": len(rows_data),
                    "note": "OpenPyXL preserves formula strings. Live calculation requires Excel COM or desktop application.",
                },
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="read_range", message=f"Error reading range: {exc}")

    def write_range(
        self,
        data: list[list[Any]],
        start_cell: str = "A1",
        sheet_name: str | None = None,
    ) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            start_row, start_col = _parse_cell_coord(start_cell)

            end_row = start_row + len(data) - 1
            max_col_len = max((len(r) for r in data), default=1)
            end_col = start_col + max_col_len - 1

            for r_idx, row_vals in enumerate(data):
                for c_idx, val in enumerate(row_vals):
                    ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)

            range_str = f"{start_cell}:{get_column_letter(end_col)}{end_row}"
            return ExcelResult(
                success=True,
                operation="write_range",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=range_str,
                message=f"Wrote {len(data)} rows to range '{range_str}'.",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="write_range", message=f"Error writing range: {exc}")

    def update_cell(self, cell: str, value: Any, sheet_name: str | None = None) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            r, c = _parse_cell_coord(cell)
            ws.cell(row=r, column=c, value=value)
            return ExcelResult(
                success=True,
                operation="update_cell",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell,
                message=f"Updated cell {cell} to '{value}'.",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="update_cell", message=f"Error updating cell: {exc}")

    def append_rows(self, rows: list[list[Any]], sheet_name: str | None = None) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            start_r = ws.max_row + 1
            for row_data in rows:
                ws.append(row_data)
            end_r = ws.max_row
            range_str = f"A{start_r}:{get_column_letter(ws.max_column)}{end_r}"
            return ExcelResult(
                success=True,
                operation="append_rows",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=range_str,
                message=f"Appended {len(rows)} rows to '{ws.title}'.",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="append_rows", message=f"Error appending rows: {exc}")

    def clear_range(self, cell_range: str, sheet_name: str | None = None) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            count = 0
            for row in ws[cell_range]:
                for cell in row:
                    cell.value = None
                    count += 1
            return ExcelResult(
                success=True,
                operation="clear_range",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell_range,
                message=f"Cleared {count} cells in range '{cell_range}'.",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="clear_range", message=f"Error clearing range: {exc}")

    def search_sheet(self, query: str, sheet_name: str | None = None, max_results: int = 50) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            q_lower = str(query).lower()
            matches = []

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and q_lower in str(cell.value).lower():
                        matches.append({
                            "cell": cell.coordinate,
                            "value": str(cell.value),
                            "row": cell.row,
                            "column": cell.column,
                        })
                        if len(matches) >= max_results:
                            break
                if len(matches) >= max_results:
                    break

            return ExcelResult(
                success=True,
                operation="search_sheet",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                message=f"Found {len(matches)} matches for '{query}'.",
                data={"query": query, "matches": matches, "count": len(matches)},
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="search_sheet", message=f"Search error: {exc}")

    def sort_range(
        self,
        cell_range: str,
        key_column: int,
        ascending: bool = True,
        has_headers: bool = True,
        sheet_name: str | None = None,
    ) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            rows = list(ws[cell_range])
            if not rows:
                return ExcelResult(success=False, operation="sort_range", message="Empty range.")

            row_vals = [[c.value for c in row] for row in rows]
            headers = []
            data_rows = row_vals

            if has_headers and len(row_vals) > 1:
                headers = row_vals[0]
                data_rows = row_vals[1:]

            idx = max(0, key_column - 1)

            def _sort_key(r):
                val = r[idx] if idx < len(r) else ""
                return (val is None, str(val).lower() if val is not None else "")

            sorted_data = sorted(data_rows, key=_sort_key, reverse=not ascending)

            final_rows = ([headers] if headers else []) + sorted_data
            for r_idx, r_data in enumerate(final_rows):
                for c_idx, val in enumerate(r_data):
                    rows[r_idx][c_idx].value = val

            return ExcelResult(
                success=True,
                operation="sort_range",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell_range,
                message=f"Sorted range '{cell_range}' by column {key_column} ({'ASC' if ascending else 'DESC'}).",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="sort_range", message=f"Sort error: {exc}")

    def filter_range(
        self,
        cell_range: str,
        column_index: int,
        criteria: str,
        sheet_name: str | None = None,
    ) -> ExcelResult:
        """Filter data non-destructively: returns matching rows and sets sheet auto-filter boundaries."""
        try:
            ws = self._get_sheet(sheet_name)
            ws.auto_filter.ref = cell_range

            rows = list(ws[cell_range])
            if not rows:
                return ExcelResult(success=False, operation="filter_range", message="Empty range.")

            row_vals = [[c.value for c in row] for row in rows]
            headers = row_vals[0] if row_vals else []
            data_rows = row_vals[1:] if len(row_vals) > 1 else []

            col_idx = max(0, column_index - 1)
            crit_lower = str(criteria).lower()

            matching_rows = []
            for r in data_rows:
                val = r[col_idx] if col_idx < len(r) else None
                if val is not None and crit_lower in str(val).lower():
                    matching_rows.append(r)

            return ExcelResult(
                success=True,
                operation="filter_range",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell_range,
                message=f"Filtered '{cell_range}': {len(matching_rows)} matching rows out of {len(data_rows)} (no rows deleted).",
                data={
                    "headers": headers,
                    "matching_rows": matching_rows,
                    "total_matches": len(matching_rows),
                    "total_data_rows": len(data_rows),
                },
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="filter_range", message=f"Filter error: {exc}")

    def write_formula(self, cell: str, formula: str, sheet_name: str | None = None) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            r, c = _parse_cell_coord(cell)
            clean_formula = formula if formula.startswith("=") else f"={formula}"
            ws.cell(row=r, column=c, value=clean_formula)
            return ExcelResult(
                success=True,
                operation="write_formula",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell,
                message=f"Wrote formula '{clean_formula}' to cell {cell}. (Note: OpenPyXL preserves formulas; live calculation requires Excel COM or opening in Excel).",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="write_formula", message=f"Formula error: {exc}")

    def format_range(
        self,
        cell_range: str,
        font_name: str | None = None,
        font_size: int | None = None,
        bold: bool | None = None,
        italic: bool | None = None,
        font_color: str | None = None,
        fill_color: str | None = None,
        number_format: str | None = None,
        alignment: str | None = None,
        borders: bool | None = None,
        auto_fit: bool = False,
        sheet_name: str | None = None,
    ) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            cell_grid = ws[cell_range] if ":" in cell_range else [[ws[cell_range]]]

            for row in cell_grid:
                for cell in row:
                    current_font = cell.font or Font()
                    font_kwargs = {
                        "name": font_name or current_font.name,
                        "size": font_size or current_font.size,
                        "bold": bold if bold is not None else current_font.bold,
                        "italic": italic if italic is not None else current_font.italic,
                    }
                    if font_color:
                        font_kwargs["color"] = _parse_color(font_color)
                    cell.font = Font(**font_kwargs)

                    if fill_color:
                        hex_c = _parse_color(fill_color)
                        cell.fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")

                    if number_format:
                        cell.number_format = number_format

                    if alignment:
                        cell.alignment = Alignment(horizontal=alignment, vertical="center")

                    if borders:
                        thin_side = Side(border_style="thin", color="000000")
                        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            if auto_fit:
                for col in ws.columns:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

            return ExcelResult(
                success=True,
                operation="format_range",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell_range,
                message=f"Formatted range '{cell_range}'.",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="format_range", message=f"Format error: {exc}")

    def create_chart(
        self,
        cell_range: str,
        chart_type: str = "col",
        title: str = "Chart",
        target_cell: str = "E2",
        sheet_name: str | None = None,
    ) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)

            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(cell_range)
            data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

            ctype = chart_type.lower()
            if ctype in ("bar", "horizontal_bar"):
                chart = BarChart()
                chart.type = "bar"
            elif ctype in ("pie", "donut"):
                chart = PieChart()
            elif ctype in ("line", "lines"):
                chart = LineChart()
            elif ctype == "scatter":
                chart = ScatterChart()
            else:
                chart = BarChart()
                chart.type = "col"

            chart.title = title
            chart.add_data(data_ref, titles_from_data=True)
            ws.add_chart(chart, target_cell)

            return ExcelResult(
                success=True,
                operation="create_chart",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                affected_range=cell_range,
                message=f"Created {chart.type or ctype} chart '{title}' placed at {target_cell}.",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="create_chart", message=f"Chart creation error: {exc}")

    def list_charts(self, sheet_name: str | None = None) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            charts = getattr(ws, "_charts", [])
            charts_info = []
            for i, c in enumerate(charts, start=1):
                t = getattr(c, "title", "")
                title_str = ""
                if isinstance(t, str):
                    title_str = t
                elif t:
                    try:
                        if hasattr(t, "tx") and t.tx and hasattr(t.tx, "rich") and t.tx.rich:
                            texts = [r.t for p in t.tx.rich.paragraphs for r in getattr(p, "r", []) if getattr(r, "t", None)]
                            title_str = "".join(texts)
                    except Exception:
                        pass
                charts_info.append({
                    "index": i,
                    "title": title_str,
                })
            return ExcelResult(
                success=True,
                operation="list_charts",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                data={"charts": charts_info, "count": len(charts_info)},
                message=f"Found {len(charts_info)} charts in sheet '{ws.title}'.",
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="list_charts", message=f"Chart error: {exc}")

    def update_chart(
        self,
        chart_identifier: str | int = 1,
        title: str | None = None,
        chart_type: str | None = None,
        cell_range: str | None = None,
        name: str | None = None,
        sheet_name: str | None = None,
    ) -> ExcelResult:
        try:
            ws = self._get_sheet(sheet_name)
            charts = getattr(ws, "_charts", [])
            if not charts:
                return ExcelResult(success=False, operation="update_chart", message=f"No charts found in sheet '{ws.title}'.")
            idx = 0
            if isinstance(chart_identifier, int) or (isinstance(chart_identifier, str) and str(chart_identifier).isdigit()):
                idx = max(0, min(int(chart_identifier) - 1, len(charts) - 1))
            chart = charts[idx]
            if title is not None:
                chart.title = title
            return ExcelResult(
                success=True,
                operation="update_chart",
                workbook=os.path.basename(self.file_path),
                sheet=ws.title,
                message=f"Updated chart title to '{title}'.",
                data={"title": str(title)},
            )
        except Exception as exc:
            return ExcelResult(success=False, operation="update_chart", message=f"Chart update error: {exc}")
